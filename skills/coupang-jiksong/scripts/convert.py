#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
쿠팡 물류동봉문서 PDF → 사방넷 발주등록(직송) 엑셀 변환
사용법:
  python3 convert.py <물류동봉문서.pdf> [<물류동봉문서2.pdf> ...] \
      [--goods <공산품상품코드.xlsx>] [--outdir <출력폴더>]
"""
import sys, os, re, json, argparse, datetime
import pdfplumber
from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
RULES = json.load(open(os.path.join(HERE, "mapping_rules.json"), encoding="utf-8"))
BUNDLE = set(RULES["bundle_skus"])
ONEPACK = RULES["onepack_rules"]
SKU_MASTER = RULES.get("sku_master", {})  # 사방넷 판매상품 마스터: 코드 → 판매상품명

COLS = ['상품고유코드','판매상품명','수량','배송방식','주문자 이름','받는분 이름',
        '전화번호1','전화번호2','우편번호','주소1','주소2','배송메세지','주문번호',
        '관리메모1','관리메모2','관리메모3','관리메모4','관리메모5',
        '상품별 메모1','상품별 메모2','상품별 메모3','발주 타입','출고희망일']

def grab(text, label):
    """'label 값' 형태 한 줄에서 값 추출."""
    for line in text.splitlines():
        line = line.strip()
        if line.startswith(label):
            return line[len(label):].strip()
    return ""

def parse_pdf(path):
    """모든 페이지 텍스트를 모아 상품 블록을 파싱. 보관용 중복 페이지는 상품번호로 dedupe."""
    import pdfplumber
    texts=[]
    info=None
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            t = pg.extract_text() or ""
            texts.append(t)
            if info is None and "업체명" in t:
                info = {
                    "업체명": grab(t, "업체명"),
                    "발주번호": grab(t, "발주번호"),
                    "운송타입": grab(t, "운송타입"),
                    "도착예정일": re.sub(r"\D", "", grab(t, "물류센터 도착예정일")),
                    "납품센터": grab(t, "납품 센터"),
                    "납품센터주소": grab(t, "납품 센터주소"),
                }
    info = info or {}
    items=[]
    seen=set()
    for t in texts:
        for it in parse_items_block(t):
            key=it.get("상품번호")
            if key in seen:
                continue
            seen.add(key)
            items.append(it)
    return info, items

ANCHOR   = re.compile(r"^(\d{1,2})\s+(\d{6,9})\s+(.*?)(\d{1,7})\s+(\d{1,7})\s*$")
BARCODE  = re.compile(r"(\d{13})")
YDATE    = re.compile(r"\bY?\s*\d{8}\b")
ONLY_DATE= re.compile(r"^\s*Y?\s*\d{8}\s*$")

def _clean_name(frag):
    frag = YDATE.sub(" ", frag)                       # 제조/소비기한 Y날짜 제거
    frag = re.sub(r"\s*\d{1,3}\s*박스\s*", " ", frag)  # N박스 제거
    frag = re.sub(r"\s+", " ", frag).strip()
    return frag

def _is_kor(s):
    return bool(re.search(r"[가-힣]", s))

def parse_items_block(text):
    """한 페이지에서 상품 블록 파싱. 두 레이아웃 모두 지원:
       (A) 페이지1형: 'No 상품번호 발주 확정' 단독, 상품명은 위/아래 줄.
       (B) 페이지2형(연속): 'No 상품번호 [상품명꼬리] 발주 확정' 한 줄에 혼재.
       앵커줄 가운데 텍스트(group3)도 상품명으로 합친다."""
    lines=[l.rstrip() for l in text.splitlines()]
    # 상품 영역 시작: 'Box 바코드' 헤더 있으면 그 다음, 없으면(연속페이지) 0
    start=0
    for i,l in enumerate(lines):
        if l.strip().startswith("Box 바코드"):
            start=i+1; break
    end=len(lines)
    for i,l in enumerate(lines):
        if l.strip().startswith("합계") or l.strip().startswith("(협력사"):
            end=i; break
    seg=lines[start:end]

    anchors=[i for i,l in enumerate(seg) if ANCHOR.match(l.strip())]
    items=[]
    for ai,a in enumerate(anchors):
        m=ANCHOR.match(seg[a].strip())
        no,prod,mid,oju,hwak=m.groups()
        # 바코드: 앵커 이후 첫 13자리
        bc=""; b_idx=None
        for j in range(a+1,len(seg)):
            mb=BARCODE.search(seg[j])
            if mb: bc=mb.group(1); b_idx=j; break
        # pre 이름: 이전 앵커의 바코드 다음 ~ 현재 앵커 전, 한글 줄(날짜단독 줄 제외)
        prev_b=-1
        if ai>0:
            pa=anchors[ai-1]
            for j in range(pa+1,a):
                if BARCODE.search(seg[j]): prev_b=j
            pre_start=max(prev_b+1, pa+1)
        else:
            pre_start=0
        pre=[seg[j] for j in range(pre_start,a) if _is_kor(seg[j]) and not ONLY_DATE.match(seg[j].strip())]
        # post 이름: 앵커+1 ~ 바코드 전, 한글 줄(날짜단독 제외)
        post=[]
        if b_idx is not None:
            post=[seg[j] for j in range(a+1,b_idx) if _is_kor(seg[j]) and not ONLY_DATE.match(seg[j].strip())]
        name=_clean_name(" ".join(pre+[mid]+post))
        items.append({"상품번호":prod,"발주수량":int(oju),"확정수량":int(hwak),
                      "바코드":bc,"상품명":name})
    return items

def is_nutrijeong(info):
    return "뉴트리정" in (info.get("업체명","") or "")

def map_nutrijeong(item):
    """반환: (코드, 사방넷 판매상품명, 경고). 1순위 마스터, 2순위 1개입 키워드→N코드."""
    prod_no = str(item.get("상품번호","")).strip()
    name = item.get("상품명","")
    if prod_no in SKU_MASTER:
        return prod_no, SKU_MASTER[prod_no], ""
    if prod_no in BUNDLE:  # 마스터 미갱신 대비 안전망
        return prod_no, "", "번들코드가 판매상품 마스터에 없음(마스터 갱신 필요)"
    nl = re.sub(r"\s+", "", name).lower()  # PDF 줄바꿈 공백 제거 후 매칭
    # 번들(2개입 이상) 판정이 1개입 키워드보다 먼저! (사용자 규칙 2026-06-05:
    # 신상품 번들은 마스터 미등록이어도 PDF 상품번호 그대로 사용)
    # 개입 표기 두 형태 모두 인식: "/2개 300정" 및 "/ 300정 2개"(슬래시 뒤 어디든)
    mb = None
    tail = nl.split("/")[-1] if "/" in nl else ""
    m_all = re.findall(r"(\d+)개(?!입)", tail)
    if m_all:
        mb = m_all[-1] if re.match(r"^\d+개", tail) is False else m_all[0]
        mb = m_all[0] if re.match(r"^\d+개", tail) else m_all[-1]
    if mb and int(mb) >= 2 and prod_no.isdigit() and len(prod_no) == 8:
        return prod_no, "", "[신상품] 마스터 미등록 번들 → PDF 상품번호 그대로 사용 (명칭은 PDF 원문)"
    # 1개입: 품목 키워드 → N코드
    for r in ONEPACK:
        if "all" in r and all(k.lower() in nl for k in r["all"]):
            return r["code"], SKU_MASTER.get(r["code"], ""), ""
    for r in ONEPACK:
        if "any" in r and any(k.lower() in nl for k in r["any"]):
            return r["code"], SKU_MASTER.get(r["code"], ""), ""
    return "???", "", "매핑실패: 신상품 → 사방넷 등록 코드 필요"

def load_goods(path):
    """바코드→(상품코드,출고상품명) 매핑.
    두 양식 지원:
      - 공산품 상품코드 엑셀: C열=상품코드, D열=출고상품명, F열=바코드
      - 재고조회 엑셀: 헤더에 '상품코드','출고상품명','바코드' 컬럼명 존재
    헤더명이 있으면 이름 기준, 없으면 위치(C/D/F) 기준."""
    wb = load_workbook(path, data_only=True); ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    def find(*names):
        for nm in names:
            if nm in header:
                return header.index(nm)
        return None
    ci = find("상품코드"); di = find("출고상품명","판매상품명"); fi = find("바코드")
    if ci is None or fi is None:
        ci, di, fi = 2, 3, 5   # 위치 기준 fallback
    bmap = {}
    for row in rows[1:]:
        if not row or len(row) <= max(ci, fi):
            continue
        bc = row[fi]
        if bc is None:
            continue
        code = row[ci]; nm = row[di] if (di is not None and len(row) > di) else None
        bmap[str(bc).strip()] = (str(code).strip() if code else "",
                                 str(nm).strip() if nm else "")
    return bmap

def map_gongsanpum(item, bmap):
    bc = str(item.get("바코드","")).strip()
    if bmap and bc in bmap:
        code, nm = bmap[bc]
        return code, nm, ""
    return "???", "", "매핑실패(공산품 바코드 미발견)"

def build_rows(info, items, goods_map):
    nutri = is_nutrijeong(info)
    rows = []
    warnings = []
    for it in items:
        if nutri:
            code, sname, warn = map_nutrijeong(it)
        else:
            code, sname, warn = map_gongsanpum(it, goods_map)
        qty = it.get("확정수량") or it.get("발주수량")
        row = {c: "" for c in COLS}
        row["상품고유코드"] = code
        row["판매상품명"] = sname or it.get("상품명","")  # 사방넷 명칭 우선, 없으면 PDF명
        row["수량"] = qty
        row["배송방식"] = "직송"
        row["받는분 이름"] = info.get("납품센터","")
        row["주소1"] = info.get("납품센터주소","")
        row["발주 타입"] = "쿠팡 - 파렛트출고"
        row["출고희망일"] = info.get("도착예정일","")
        rows.append(row)
        if warn:
            warnings.append(f"  [검수필요] {it.get('상품명','')} (상품번호 {it.get('상품번호','')}, 바코드 {it.get('바코드','')}) → {warn}")
    account = "뉴트리정" if nutri else "이더컴퍼니"
    return account, rows, warnings

def write_xlsx(account, rows, info, outdir):
    wb = Workbook(); ws = wb.active; ws.title = "발주등록_sample"
    ws.append(COLS)
    for r in rows:
        ws.append([r[c] for c in COLS])
    # 출고희망일/코드 문자열 유지
    for col_idx, c in enumerate(COLS, start=1):
        if c in ("출고희망일","상품고유코드","우편번호","전화번호1"):
            for rr in range(2, ws.max_row+1):
                cell = ws.cell(row=rr, column=col_idx)
                if cell.value not in (None, ""):
                    cell.value = str(cell.value)
                    cell.number_format = "@"
    ts = datetime.datetime.now().strftime("%Y%m%d")
    order = info.get("발주번호","") or ts
    fn = f"사방넷_{account}_직송_{ts}_{order}.xlsx"
    out = os.path.join(outdir, fn)
    wb.save(out)
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("pdfs", nargs="+")
    ap.add_argument("--goods", default=None)
    ap.add_argument("--outdir", default=".")
    a = ap.parse_args()
    goods_map = load_goods(a.goods) if a.goods else {}
    os.makedirs(a.outdir, exist_ok=True)
    # 계정별로 누적
    bucket = {}   # account -> rows
    meta = {}     # account -> info (대표)
    all_warn = []
    for p in a.pdfs:
        info, items = parse_pdf(p)
        account, rows, warns = build_rows(info, items, goods_map)
        bucket.setdefault(account, []).extend(rows)
        meta.setdefault(account, info)
        print(f"[{os.path.basename(p)}] 업체={info.get('업체명')} 발주번호={info.get('발주번호')} "
              f"도착예정일={info.get('도착예정일')} 센터={info.get('납품센터')} 상품={len(items)}건")
        for it in items:
            print(f"    - {it.get('상품명','')[:40]} | 상품번호 {it.get('상품번호')} | 수량 {it.get('확정수량') or it.get('발주수량')}")
        all_warn += warns
    outs = []
    for account, rows in bucket.items():
        out = write_xlsx(account, rows, meta[account], a.outdir)
        outs.append(out)
        print(f"==> 생성: {out} ({len(rows)}행)")
    if all_warn:
        print("\n[!] 검수 필요 항목:")
        print("\n".join(all_warn))
    else:
        print("\n[OK] 매핑 실패 없음")
    return outs

if __name__ == "__main__":
    main()
